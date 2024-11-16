import openpyxl
import csv
import os
from datetime import datetime

def convert_xlsm_to_csv(input_file, output_file):
    # Verifica se o arquivo de entrada existe
    if not os.path.exists(input_file):
        print(f"Erro: O arquivo {input_file} não foi encontrado.")
        return

    try:
        # Abrir o arquivo XLSX ou XLSM usando openpyxl
        wb = openpyxl.load_workbook(input_file, data_only=True)
        print(f"Arquivo {input_file} aberto com sucesso.")
    except Exception as e:
        print(f"Erro ao abrir o arquivo XLSX/XLSM: {e}")
        return

    # Seleciona a primeira planilha
    sheet = wb.active

    if sheet.max_row == 0:
        print(f"Aviso: A planilha {input_file} está vazia.")
        return

    rows_written = 0

    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as csv_file:  # Usando UTF-8
            writer = csv.writer(csv_file)

            # Itera sobre as linhas da planilha
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                new_row = []
                for cell in row:
                    # Se a célula for numérica, mantenha como número ou substitua por 0
                    if isinstance(cell.value, (int, float)):
                        cell_value = str(cell.value) if cell.value is not None else "0"  # Coloca 0 se o valor for None
                    elif isinstance(cell.value, datetime):  # Se for uma data
                        cell_value = cell.value.strftime("%d/%m/%Y")  # Formata a data como 'dd/MM/yyyy'
                    elif cell.value is None or str(cell.value).strip() == "":  # Células vazias
                        cell_value = ""  # Deixe vazio para células vazias
                    else:  # Caso contrário, é texto
                        cell_value = str(cell.value).strip()  # Remove espaços extras
                    # Adiciona a célula processada à linha
                    new_row.append(cell_value.replace("\n", " ").replace("\r", ""))  # Substitui quebras de linha
                writer.writerow(new_row)
                rows_written += 1

        if rows_written == 0:
            print(f"Nenhum dado foi escrito no arquivo CSV. O arquivo pode estar vazio.")
        else:
            print(f"{rows_written} linhas escritas no arquivo CSV.")
    except Exception as e:
        print(f"Erro ao salvar o arquivo CSV: {e}")
